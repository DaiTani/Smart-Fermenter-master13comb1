# deepseek_combined.py
import concurrent.futures
import minimalmodbus
import time
import keyboard
import serial
from serial import SerialException
import threading
import queue
import csv
import os
from datetime import datetime, timedelta
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
from matplotlib.animation import FuncAnimation
import matplotlib.dates as mdates
import cProfile
import pstats
import io
import serial.tools.list_ports  # 新增：用于列出可用串口
import json  # 新增：用于读取OCR配置文件
import pytesseract  # 新增：用于OCR识别
from PIL import ImageGrab, Image  # 新增：用于截图
import pygetwindow as gw  # 新增：用于窗口操作
import numpy as np  # 新增：用于图像处理
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# 设置Tesseract路径（根据您的实际安装路径修改）
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
# ------------------ 全局配置 -------------------
# MODBUS配置（泵控制）
PORT_MODBUS = 'COM9'  # 修改为独立的阀门控制串口
BAUDRATE_MODBUS = 9600
SLAVE_ADDRESS = 1

# 串口配置（双数据采集）
COM3_CONFIG = {    #电导
    'port': 'COM7',
    'baudrate': 115200,
    'timeout': 0.1  # 已优化为0.1秒
}

COM6_CONFIG = {         #荧光
    'port': 'COM10',
    'baudrate': 9600,
    'bytesize': 8,
    'parity': serial.PARITY_NONE,
    'stopbits': 1,
    'slave_address': 1
}

COM7_CONFIG = {        #浊度1
    'port': 'COM4',
    'baudrate': 9600,
    'bytesize': 8,
    'parity': serial.PARITY_NONE,
    'stopbits': 1,
    'timeout': 1
}

# 新增：COM15浊度传感器配置（浊度2）
COM15_CONFIG = {        #浊度2
    'port': 'COM6',
    'baudrate': 9600,
    'bytesize': 8,
    'parity': serial.PARITY_NONE,
    'stopbits': 1,
    'timeout': 1
}

# 新增：OCR配置文件路径
OCR_CONFIG_PATH = "ocr_config.json"  # OCR配置文件路径

AD_RESOLUTION = 4096
VOLTAGE_REF = 10.0

CMD_REGISTER = 0
POS_REGISTER = 1
SPD_REGISTER = 3
STATUS_REGISTER = 4
VALVE_CHANNEL_CMD = 0x08

MAX_STEPS = 384000
MAX_SPEED = 50000
VALVE_CHANNELS = [1]
#VALVE_CHANNELS = [1, 1, 1, 1]

command_queue = queue.Queue()
data_queue = queue.Queue(maxsize=1000)  # 限制队列大小
stop_event = threading.Event()


# ------------------ 新泵控制模块 -------------------
class NewPumpController:
    def __init__(self, port, pump_address):
        self.ser = None
        self.port = port
        self.pump_address = pump_address
        self.RUN_REGISTER = 0x0000
        self.DIRECTION_REGISTER = 0x0001
        self.SPEED_REGISTER = 0x0002
        #############################self.initialize_serial()
        
    def initialize_serial(self):
        max_attempts = 3
        for attempt in range(max_attempts):
            try:
                self.ser = serial.Serial(
                    port=self.port,
                    baudrate=9600,
                    bytesize=serial.EIGHTBITS,
                    stopbits=serial.STOPBITS_ONE,  # 修改为1个停止位（更常见）
                    parity=serial.PARITY_NONE,
                    timeout=1
                )
                print(f"{self.port} 泵串口连接成功")
                return
            except SerialException as e:
                ########################################print(f"{self.port} 泵连接失败 (尝试 {attempt+1}/{max_attempts}): {e}")
                time.sleep(1)
        
        print(f"错误：无法连接 {self.port}，程序退出")
        exit(1)
    
    def calculate_crc(self, data):
        """计算 Modbus RTU CRC 校验码（低字节在前）"""
        crc = 0xFFFF
        for byte in data:
            crc ^= byte
            for _ in range(8):
                if crc & 0x0001:
                    crc >>= 1
                    crc ^= 0xA001  # 多项式 x^16 + x^15 + x^2 + 1
                else:
                    crc >>= 1
        return crc

    def write_register(self, reg_addr, reg_value):
        """
        写入单个寄存器（功能码 0x06）
        返回 True 表示成功，False 表示失败
        """
        function_code = 0x06
        # 构建请求帧
        data = bytes([self.pump_address, function_code]) + reg_addr.to_bytes(2, 'big') + reg_value.to_bytes(2, 'big')
        crc = self.calculate_crc(data)
        crc_bytes = [(crc & 0xFF), (crc >> 8) & 0xFF]  # CRC 低字节在前
        frame = data + bytes(crc_bytes)
        
        # 发送请求
        self.ser.write(frame)
        time.sleep(0.2)  # 增加等待时间
        
        # 读取响应（最大读取长度为 100 字节）
        response = self.ser.read(100)
        if not response:
            print(f"{self.port} 泵未收到响应")
            return False
        
        # 响应长度检查（功能码0x06的响应应为8字节）
        if len(response) < 8:
            print(f"{self.port} 泵响应长度不足: {len(response)}")
            return False
        
        # 解析响应
        resp_address = response[0]
        resp_func = response[1]
        resp_reg = int.from_bytes(response[2:4], 'big')
        resp_val = int.from_bytes(response[4:6], 'big')
        received_crc = int.from_bytes(response[6:8], 'little')  # CRC 低字节在前
        
        # 重新计算 CRC
        expected_crc = self.calculate_crc(response[:6])
        expected_crc_low = expected_crc & 0xFF
        expected_crc_high = (expected_crc >> 8) & 0xFF
        
        # 验证响应内容
        if (resp_address == self.pump_address and 
            resp_func == function_code and 
            resp_reg == reg_addr and 
            resp_val == reg_value and 
            received_crc == ((expected_crc_high << 8) | expected_crc_low)):
            return True
        else:
            print(f"{self.port} 泵响应校验失败")
            return False

    def start_pump(self, direction, speed_rpm):
        """启动泵，设置方向和速度"""
        # 设置方向
        if not self.write_register(self.DIRECTION_REGISTER, direction):
            print(f"{self.port} 泵方向设置失败")
            return False
        
        # 设置速度（速度值 = RPM * 10）
        speed_value = int(speed_rpm * 10)
        if not self.write_register(self.SPEED_REGISTER, speed_value):
            print(f"{self.port} 泵速度设置失败")
            return False
        
        # 启动泵
        if self.write_register(self.RUN_REGISTER, 0x0001):
            print(f"{self.port} 泵已启动 - {'正转' if direction == 0x0001 else '反转'}，速度{speed_rpm}转/分钟")
            return True
        return False

    def stop_pump(self):
        """停止泵"""
        if self.write_register(self.RUN_REGISTER, 0x0000):
            print(f"{self.port} 泵已停止")
            return True
        print(f"{self.port} 泵停止失败")
        return False

    def safe_shutdown(self):
        """安全关闭泵"""
        stop_success = False
        for attempt in range(3):
            print(f"尝试停止 {self.port} 泵... (尝试 {attempt+1}/3)")
            if self.stop_pump():
                stop_success = True
                break
            time.sleep(0.5)
        
        if not stop_success:
            print(f"警告：未能确认 {self.port} 泵停止状态！")
        return stop_success

# ------------------ 阀门控制模块 -------------------
class ValveController:
    def __init__(self):
        self.instrument = None
        self.initialize_modbus()
        
    def initialize_modbus(self):
        try:
            #self.instrument = minimalmodbus.Instrument(PORT_MODBUS, SLAVE_ADDRESS)
            #self.instrument.serial.baudrate = BAUDRATE_MODBUS
            #self.instrument.serial.timeout = 1
            #self.instrument.mode = minimalmodbus.MODE_RTU
            #self.instrument.clear_buffers_before_each_transaction = True
            print("阀门控制器初始化成功")
        except SerialException as e:
            print(f"阀门控制器连接失败: {e}")
            exit()

    def check_status(self):
        try:
            response = self.instrument.read_registers(STATUS_REGISTER, 2, functioncode=4)
            status_low = response[0]
            status_high = response[1]
            motor_busy = (status_low >> 8) & 0x01 == 0
            valve_channel = status_high & 0x0F
            return {"motor_busy": motor_busy, "valve_channel": valve_channel}
        except Exception as e:
            print(f"阀门状态读取失败: {e}")
            return {"motor_busy": True, "valve_channel": 0}

    def send_command(self, command, parameter=0, wait=True):
        try:
            value = (command << 8) | parameter
            self.instrument.write_register(CMD_REGISTER, value, functioncode=6)
            if wait:
                while self.check_status()["motor_busy"]:
                    time.sleep(0.1)
            print(f"阀门指令 0x{command:02X} 参数 0x{parameter:02X} 发送成功")
            return True
        except Exception as e:
            print(f"阀门指令发送失败: {e}")
            return False

    def switch_valve(self, channel):
        if 1 <= channel <= 15:
            print(f"\n切换到通道 {channel}...")
            command_queue.put(0)
            success = self.send_command(VALVE_CHANNEL_CMD, channel, wait=True)
            if success:
                time.sleep(3)
            return success
        print("\n无效的通道号（1-15）")
        return False

    def initialize_valve(self):
        print("\n正在初始化阀门控制器...")
        return self.send_command(0x06, 0x01)
        
    def check_pause(self):
        if keyboard.is_pressed('shift') or keyboard.is_pressed('left shift') or keyboard.is_pressed('right shift'):
            print("\n检测到Shift键，程序暂停...（按回车继续）")
            start_pause = time.time()
            while True:
                if keyboard.is_pressed('enter'):
                    print("继续运行...")
                    return True
                if stop_event.is_set():
                    return False
                if time.time() - start_pause > 3600:
                    print("暂停超时，继续运行...")
                    return True
                time.sleep(0.05)
        return True

# ------------------ 主控制模块 -------------------
class MainController:
    def __init__(self):
        self.valve_controller = ValveController()
        self.pump1 = NewPumpController('COM12', 0x01)  # 1号泵
        self.pump2 = NewPumpController('COM13', 0x01)  # 2号泵
        
    def initialize_system(self):
        if not self.valve_controller.initialize_valve():
            return False
        return True
    
    def main_loop(self):
        try:
            #if not self.initialize_system():
            #    return
                
            while not stop_event.is_set() and not keyboard.is_pressed('esc'):
                for channel in VALVE_CHANNELS:
                    print(f"\n--- 处理通道 {channel} ---")
                    
                    #if not self.valve_controller.switch_valve(channel):
                    #    continue
                    #time.sleep(2)
                    
                    print("\n执行吸液操作...")
                    #command_queue.put(0)
                    # 使用1号泵进行吸液
                    if not self.pump1.start_pump(0x0001, 20):  # 正转，20转/分钟
                        continue
                    #time.sleep(5)  # 吸液5秒
                    #self.pump1.stop_pump()
                    #time.sleep(0.5)
                    
                    #if not self.valve_controller.switch_valve(6):
                    #    continue
                    #time.sleep(2)
                    
                    #print("\n执行排液操作...")
                    #command_queue.put(1)
                    # 使用2号泵进行排液
                    if not self.pump2.start_pump(0x0001, 20):  # 正转，10转/分钟
                        continue
                    for _ in range(20):
                        if stop_event.is_set():
                            return
                        time.sleep(1)    
                    time.sleep(3600)  # 
                    #self.pump2.stop_pump()
                    #if not self.pump1.start_pump(0x0001, 0):  # 正转，10转/分钟
                    #    continue                    
                    #if not self.pump2.start_pump(0x0001, 0):  # 正转，20转/分钟
                    #    continue                    
                    
                    #for _ in range(20):
                    #    if stop_event.is_set():
                    #        return
                    #    time.sleep(1)
                    #time.sleep(1200)
                    
                    print(f"\n通道 {channel} 完成3600秒，等待运行...")
                    start_time = time.time()
                    while time.time() - start_time < 1:
                        if stop_event.is_set() or keyboard.is_pressed('esc'):
                            return
                        time.sleep(0.1)
                        
        except KeyboardInterrupt:
            print("程序手动终止")
            stop_event.set()
        finally:
            # 安全关闭所有设备
            self.pump1.safe_shutdown()
            self.pump2.safe_shutdown()
            #self.valve_controller.send_command(0x04, 0x00)
            stop_event.set()

class OCRProcessor:
    def __init__(self):
        self.config = None
        self.target_window = None
        self.ocr_regions = []
        self.last_ocr_results = {}  # 存储上一次成功的OCR结果
        self.debug_dir = "ocr_debug"  # 调试图像保存目录
        os.makedirs(self.debug_dir, exist_ok=True)  # 创建调试目录
        self.load_config()
        
    def load_config(self):
        """加载OCR配置文件"""
        try:
            if os.path.exists(OCR_CONFIG_PATH):
                with open(OCR_CONFIG_PATH, 'r', encoding='utf-8') as f:
                    self.config = json.load(f)
                    self.target_window = self.config.get("window_title", "")
                    self.ocr_regions = self.config.get("regions", [])
                    print(f"OCR配置已加载: {len(self.ocr_regions)}个区域")
                    
                    # 初始化上一次OCR结果
                    for region in self.ocr_regions:
                        region_name = region.get("name", "")
                        if region_name:
                            self.last_ocr_results[region_name] = ""
            else:
                print("未找到OCR配置文件")
        except Exception as e:
            print(f"加载OCR配置失败: {str(e)}")
    
    def process_single_region(self, region, screenshot, timestamp):
        """处理单个区域的OCR识别"""
        region_name = region.get("name", "")
        if not region_name:
            return region_name, ""
        
        rect = region.get("rect", region)
        x = rect.get("x", 0)
        y = rect.get("y", 0)
        width = rect.get("width", 0)
        height = rect.get("height", 0)
        
        # 裁剪区域图像
        try:
            region_img = screenshot.crop((x, y, x + width, y + height))
            
            # 保存原始区域图像（调试用）
            #region_img.save(os.path.join(self.debug_dir, f"{timestamp}_{region_name}_original.png"))
        except Exception as e:
            print(f"裁剪区域 '{region_name}' 失败: {str(e)}")
            return region_name, self.last_ocr_results.get(region_name, "")
        
        # 优化图像预处理
        img = region_img.resize((region_img.width // 2, region_img.height // 2))
        img = region_img.convert('L')  # 转为灰度
        
        # 二值化处理 - 使用自适应阈值
        img = img.point(lambda p: 0 if p < 180 else 255)  # 提高阈值
        
        # 保存预处理后的图像（调试用）
        #img.save(os.path.join(self.debug_dir, f"{timestamp}_{region_name}_processed.png"))
        
        # 使用Tesseract进行OCR识别
        try:
            # 尝试多种OCR配置
            configs = [
                #'--psm 6 digits',  # 假设为统一文本块
                #'--psm 7 digits',  # 单行文本
                #'--psm 8 digits',  # 单个单词
                '--psm 10 digits'  # 单个字符
            ]
            
            best_text = ""
            for config in configs:
                text = pytesseract.image_to_string(img, config=config)
                
                # 清理识别结果
                cleaned_text = ''.join(filter(lambda x: x.isdigit() or x in ['.', '-'], text))
                
                # 检查是否为有效的数字格式
                if cleaned_text and any(char.isdigit() for char in cleaned_text):
                    if not best_text:  # 第一个有效结果
                        best_text = cleaned_text
                    elif len(cleaned_text) > len(best_text):  # 选择最长的有效结果
                        best_text = cleaned_text
            
            if best_text:
                print(f"区域 '{region_name}' 识别成功: {best_text}")
                return region_name, best_text
            else:
                print(f"警告：区域 '{region_name}' 识别失败，使用上一次的值")
                return region_name, self.last_ocr_results.get(region_name, "")
        except Exception as e:
            print(f"OCR识别出错: {str(e)}")
            return region_name, self.last_ocr_results.get(region_name, "")
    
    def perform_ocr(self):
        """执行OCR识别并返回结果，失败时使用上一次成功的结果"""
        results = {}
        if not self.target_window or not self.ocr_regions:
            print("OCR配置不完整，跳过识别")
            return results
            
        try:
            # 获取目标窗口
            windows = gw.getWindowsWithTitle(self.target_window)
            if not windows:
                print(f"未找到窗口: {self.target_window}")
                # 返回上一次的结果
                return self.last_ocr_results
                
            window = windows[0]
            wx, wy = window.left, window.top
            
            # 激活窗口并置顶
            # 激活窗口并置顶
            try:
                window.activate()
                window.restore()
                window.bringToFront()
            except Exception:
                pass
            time.sleep(0.01)
            
            # 只截取一次整个窗口的图像
            try:
                screenshot = ImageGrab.grab(bbox=(wx, wy, wx + window.width, wy + window.height))
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                
                # 保存整个窗口图像（调试用）
                #screenshot.save(os.path.join(self.debug_dir, f"{timestamp}_full_window.png"))
            except Exception as e:
                print(f"截取窗口图像失败: {str(e)}")
                return self.last_ocr_results
            
            # 使用线程池并行处理多个区域
            with concurrent.futures.ThreadPoolExecutor() as executor:
                # 为每个区域创建处理任务
                futures = []
                for region in self.ocr_regions:
                    # 复制区域坐标，避免多线程冲突
                    region_copy = region.copy()
                    futures.append(executor.submit(self.process_single_region, region_copy, screenshot, timestamp))
                
                # 获取所有结果
                for future in concurrent.futures.as_completed(futures):
                    region_name, result = future.result()
                    if region_name:
                        results[region_name] = result
                        # 更新上一次成功的结果
                        self.last_ocr_results[region_name] = result
            
            return results
            
        except Exception as e:
            print(f"执行OCR识别时出错: {str(e)}")
            # 出错时返回上一次的结果
            return self.last_ocr_results

class KalmanFilter:
    def __init__(self, process_variance=1e-4, measurement_variance=0.1, initial_value=0):
        """
        初始化卡尔曼滤波器
        :param process_variance: 过程噪声方差（系统不确定性）
        :param measurement_variance: 测量噪声方差（传感器不确定性）
        :param initial_value: 初始值
        """
        self.process_variance = process_variance
        self.measurement_variance = measurement_variance
        self.posteri_estimate = initial_value
        self.posteri_error_estimate = 1.0  # 初始估计误差
    
    def update(self, measurement):
        """
        更新滤波器状态
        :param measurement: 新的测量值
        :return: 滤波后的值
        """
        # 预测步骤
        priori_estimate = self.posteri_estimate
        priori_error_estimate = self.posteri_error_estimate + self.process_variance
        
        # 更新步骤
        kalman_gain = priori_error_estimate / (priori_error_estimate + self.measurement_variance)
        self.posteri_estimate = priori_estimate + kalman_gain * (measurement - priori_estimate)
        self.posteri_error_estimate = (1 - kalman_gain) * priori_error_estimate
        
        return self.posteri_estimate            
# ------------------ 数据采集模块 -------------------
class DataCollector:
    def __init__(self):
        self.ser_com3 = None
        self.instrument_com6 = None
        self.ser_com7 = None
        self.ser_com15 = None  # 新增：COM15浊度传感器
        self.latest_com3_data = [0.0, 0.0]
        self.latest_com6_data = [0.0, 0.0]
        self.latest_com7_data = [0, 0]  # 浊度1数据
        self.latest_com15_data = [0, 0]  # 新增：浊度2数据
        self.data_lock = threading.Lock()
        self.ocr_processor = OCRProcessor()  # 新增：OCR处理器
        #################self.initialize_serial()
        self.last_log = ""
        self.channel6_files = {}
        self.current_normal_file = None  # 新增：单独管理普通通道文件
        self.kalman_filters = {}
        self.initialize_kalman_filters()        
        self.sensor_buffer = [[] for _ in range(8)]  # 8个传感器通道
        self.buffer_size = 5  # 缓冲10个数据点
        self.buffer_count = 0        
        
        
    def initialize_serial(self):
        # COM3初始化
        try:
            self.ser_com3 = serial.Serial(**COM3_CONFIG)
        except SerialException as e:
            #################print(f"COM3连接失败: {e}")
            exit()

        # COM6初始化
        try:
            self.instrument_com6 = minimalmodbus.Instrument(
                COM6_CONFIG['port'], 
                COM6_CONFIG['slave_address']
            )
            self.instrument_com6.serial.baudrate = COM6_CONFIG['baudrate']
            self.instrument_com6.serial.bytesize = COM6_CONFIG['bytesize']
            self.instrument_com6.serial.parity = COM6_CONFIG['parity']
            self.instrument_com6.serial.stopbits = COM6_CONFIG['stopbits']
            self.instrument_com6.serial.timeout = 1
            self.instrument_com6.mode = minimalmodbus.MODE_RTU
            self.instrument_com6.clear_buffers_before_each_transaction = True
        except Exception as e:
            #################print(f"COM6连接失败: {e}")
            exit()

        # COM7初始化
        try:
            self.ser_com7 = serial.Serial(**COM7_CONFIG)
            self.ser_com7.reset_input_buffer()
        except SerialException as e:
            #################print(f"COM7连接失败: {e}")
            exit()
            
        # 新增：COM15初始化
        try:
            self.ser_com15 = serial.Serial(**COM15_CONFIG)
            self.ser_com15.reset_input_buffer()
            print("COM15浊度传感器连接成功")
        except SerialException as e:
            #################print(f"COM15浊度传感器连接失败: {e}")
            # 即使连接失败也要继续运行，但使用默认值
            self.ser_com15 = None

    def read_com7_data(self):
        try:
            dirt_value, ad_value = None, None
            cmd_dirt = bytes([0x18, 0x05, 0x00, 0x01, 0x0D])
            self.ser_com7.write(cmd_dirt)
            time.sleep(0.05)  # 优化后的延迟
            response = self.ser_com7.read(5)
            if len(response) == 5 and response[0] == 0x18 and response[-1] == 0x0D:
                dirt_value = response[3]

            cmd_ad = bytes([0x18, 0x05, 0x00, 0x02, 0x0D])
            self.ser_com7.write(cmd_ad)
            time.sleep(0.05)  # 优化后的延迟
            response_ad = self.ser_com7.read(6)
            if len(response_ad) == 6 and response_ad[0] == 0x18 and response_ad[-1] == 0x0D:
                ad_high = response_ad[3]
                ad_low = response_ad[4]
                ad_value = (ad_high << 8) | ad_low

            return (dirt_value if dirt_value is not None else self.latest_com7_data[0], 
                    ad_value if ad_value is not None else self.latest_com7_data[1])
        except Exception as e:
            #################print(f"COM7数据读取失败: {e}")
            return (self.latest_com7_data[0], self.latest_com7_data[1])
            
    # 新增：读取COM15浊度传感器数据（与COM7相同）
    def read_com15_data(self):
        try:
            if not self.ser_com15:
                return (self.latest_com15_data[0], self.latest_com15_data[1])
                
            dirt_value, ad_value = None, None
            cmd_dirt = bytes([0x18, 0x05, 0x00, 0x01, 0x0D])
            self.ser_com15.write(cmd_dirt)
            time.sleep(0.05)
            response = self.ser_com15.read(5)
            if len(response) == 5 and response[0] == 0x18 and response[-1] == 0x0D:
                dirt_value = response[3]

            cmd_ad = bytes([0x18, 0x05, 0x00, 0x02, 0x0D])
            self.ser_com15.write(cmd_ad)
            time.sleep(0.05)
            response_ad = self.ser_com15.read(6)
            if len(response_ad) == 6 and response_ad[0] == 0x18 and response_ad[-1] == 0x0D:
                ad_high = response_ad[3]
                ad_low = response_ad[4]
                ad_value = (ad_high << 8) | ad_low

            return (dirt_value if dirt_value is not None else self.latest_com15_data[0], 
                    ad_value if ad_value is not None else self.latest_com15_data[1])
        except Exception as e:
            print(f"COM15数据读取失败: {e}")
            return (self.latest_com15_data[0], self.latest_com15_data[1])

    def read_com3_data(self):
        try:
            data = self.ser_com3.read_until(b'$').decode('utf-8', 'ignore').strip()
            if not data.startswith('#') or '$' not in data:
                return None
            content = data[len('#'):data.index('$')]
            parts = content.split(',')
            num_list = []
            for part in parts[:2]:
                try:
                    num = float(part)
                    num_list.append(num)
                except ValueError:
                    num_list.append(-32768)
            return num_list
        except Exception as e:
            #################print(f"COM3数据解析失败: {e}")
            return None

    def read_com6_data(self):
        try:
            raw_values = self.instrument_com6.read_registers(
                registeraddress=0x0000,
                number_of_registers=2,
                functioncode=3
            )
            voltage1 = (raw_values[0] / AD_RESOLUTION) * VOLTAGE_REF
            return [raw_values[0], round(voltage1, 3)]
        except Exception as e:
            #################print(f"COM6数据读取失败: {e}")
            return None

    def perform_ocr_and_get_data(self):
        """执行OCR并获取数据"""
        ocr_data = self.ocr_processor.perform_ocr()
        return ocr_data            
            

    def preprocess_all_data(self, data):
        """预处理所有数据，将0值替换为0.0001"""
        preprocessed = []
        for value in data:
            # 尝试将值转换为浮点数
            try:
                num_value = float(value)
                # 如果值为0，替换为0.0001
                if num_value == 0:
                    preprocessed.append(0.0001)
                else:
                    preprocessed.append(num_value)
            except (ValueError, TypeError):
                # 如果无法转换为浮点数，保持原样
                preprocessed.append(value)
        return preprocessed


    def initialize_kalman_filters(self):
        """为每个传感器数据创建卡尔曼滤波器"""
        # 定义每个传感器的初始参数
        # 过程噪声方差（系统不确定性）和测量噪声方差（传感器不确定性）
        filter_params = {
            # OD (COM3)
            0: {"process_variance": 1e-6, "measurement_variance": 0.5, "initial_value": 0.0},
            # electrical conductivity (COM3)
            1: {"process_variance": 1e-6, "measurement_variance": 0.5, "initial_value": 0.0},
            # fluorescence raw (COM6)
            2: {"process_variance": 1e-6, "measurement_variance": 0.5, "initial_value": 0.0},
            # fluorescence voltage (COM6)
            3: {"process_variance": 1e-6, "measurement_variance": 0.5, "initial_value": 0.0},
            # Hum%RH (COM7 dirt value)
            4: {"process_variance": 1e-6, "measurement_variance": 0.5, "initial_value": 0.0},
            # AD Value1 (COM7 ad value)
            5: {"process_variance": 1e-6, "measurement_variance": 0.5, "initial_value": 0.0},
            # Dirt2 (COM15 dirt value)
            6: {"process_variance": 1e-6, "measurement_variance": 0.5, "initial_value": 0.0},
            # AD Value2 (COM15 ad value)
            7: {"process_variance": 1e-6, "measurement_variance": 0.5, "initial_value": 0.0},
        }
        
        for idx, params in filter_params.items():
            self.kalman_filters[idx] = KalmanFilter(
                process_variance=params["process_variance"],
                measurement_variance=params["measurement_variance"],
                initial_value=params["initial_value"]
            )
    
    def apply_kalman_filter(self, data):
        """
        对传感器数据应用卡尔曼滤波
        :param data: 原始传感器数据列表 (8个元素)
        :return: 滤波后的数据列表
        """
        filtered_data = []
        for idx, value in enumerate(data):
            try:
                # 只对数值类型应用滤波
                if isinstance(value, (int, float)):
                    # 确保滤波器存在
                    if idx in self.kalman_filters:
                        filtered_value = self.kalman_filters[idx].update(value)
                        filtered_data.append(filtered_value)
                    else:
                        filtered_data.append(value)
                else:
                    filtered_data.append(value)
            except Exception as e:
                print(f"卡尔曼滤波出错 (索引 {idx}): {str(e)}")
                filtered_data.append(value)
        return filtered_data

    def add_to_buffer(self, filtered_data):
        """将滤波后的数据添加到缓冲区"""
        for i, value in enumerate(filtered_data):
            if len(self.sensor_buffer[i]) < self.buffer_size:
                self.sensor_buffer[i].append(value)
            else:
                # 缓冲区已满，替换最旧的数据
                self.sensor_buffer[i].pop(0)
                self.sensor_buffer[i].append(value)
        
        self.buffer_count += 1
        
        # 检查缓冲区是否已满
        return self.buffer_count >= self.buffer_size
    
    def process_buffer(self):
        """处理缓冲区中的数据，返回去极值平均后的数据"""
        processed_data = []
        
        for channel in self.sensor_buffer:
            if len(channel) < 3:
                # 数据不足，直接返回平均值
                if channel:
                    processed_data.append(sum(channel) / len(channel))
                else:
                    processed_data.append(0.0)
                continue
            
            # 复制数据并排序
            sorted_data = sorted(channel)
            
            # 去掉一个最小值和一个最大值
            trimmed_data = sorted_data[1:-1]
            
            # 计算平均值
            avg = sum(trimmed_data) / len(trimmed_data)
            processed_data.append(avg)
        
        # 重置缓冲区计数
        self.buffer_count = 0
        
        return processed_data

        
    def serial_to_excel(self):
        # 定义新的保存目录
        base_dir = r".\Data5\28"
        main_file_path = os.path.join(base_dir, 'data.xlsx')
        os.makedirs(base_dir, exist_ok=True)  # 确保目录存在
        last_timestamp_value = self.get_last_timestamp_value(main_file_path)# 初始化最后一个时间戳值
        pr = cProfile.Profile()
        pr.enable()
        
        # 备份逻辑保持不变，但文件扩展名改为 .xlsx 并存储在 base_dir
        backup_files = [
            os.path.join(base_dir, f'sensor_data{channel}.xlsx') for channel in range(1,5)
        ] + [
            os.path.join(base_dir, 'data.xlsx')  # 主文件重命名为 data.xlsx
        ] + [
            os.path.join(base_dir, f'sensor_data6_{c}.xlsx') for c in range(1,5)
        ]
        
        for filename in backup_files:
            if os.path.exists(filename):
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                backup_name = f"{filename.split('.')[0]}_{timestamp}.xlsx"
                os.rename(filename, backup_name)
        
        # 修改Excel表头，添加OCR数据列
        headers = ['Timestamp', 'OD', 'electrical conductivity', 'fluorescence', 
                   'Hum%RH', 'Dirt1', 'AD Value1', 'Dirt2', 'AD Value2']





                   
        # 添加OCR列标题
        if self.ocr_processor.ocr_regions:
            for region in self.ocr_processor.ocr_regions:
                region_name = region.get("name", "")
                if region_name:
                    headers.append(f'{region_name}')
        
        # 创建工作簿和工作表对象
        from openpyxl import Workbook
        
        # 主工作簿 - 存储在 base_dir 并命名为 data.xlsx
        merged_wb = Workbook()
        merged_ws = merged_wb.active
        merged_ws.append(headers + ['Log'])  # 添加日志列
        
        # 主文件路径
        main_file_path = os.path.join(base_dir, 'data.xlsx')
        
        # 通道工作簿字典
        channel_wbs = {}
        
        current_channel = None
        buffer = []
        merged_buffer = []
        n = 0
        last_save = time.time()
        
        print("\n初始化完成，等待10秒后开始采集数据...", flush=True)
        for i in range(10, 0, -1):
            if stop_event.is_set():
                print("\n检测到终止信号，程序退出。")
                return
            print(f"\r{i} 秒后开始采集...（按Ctrl+C可退出采集）", end='', flush=True)
            time.sleep(1)
        
        try:
            while not stop_event.is_set():
                if not command_queue.empty():
                    cmd = command_queue.get()
                    if isinstance(cmd, tuple):
                        if cmd[0] == 'channel':
                            new_channel = cmd[1]
                            if new_channel != current_channel:
                                # 保存缓冲区数据到当前通道文件
                                if current_channel in channel_wbs and len(buffer) > 0:
                                    ws = channel_wbs[current_channel].active
                                    for row in buffer:
                                        ws.append(row)
                                    buffer.clear()
                                
                                # 创建新通道文件（仅限普通通道）存储在 base_dir
                                file_name = f'sensor_data{new_channel}.xlsx'
                                file_path = os.path.join(base_dir, file_name)
                                if new_channel not in channel_wbs:
                                    wb = Workbook()
                                    ws = wb.active
                                    ws.append(headers)  # 只写传感器头部，不含OCR和日志
                                    channel_wbs[new_channel] = (wb, file_path)
                                current_channel = new_channel
                                print(f"\n切换到通道 {new_channel}，数据将保存至 {file_path}")

                        elif cmd[0] == 'channel6':
                            parent_channel = cmd[1]
                            file_name = f'sensor_data6_{parent_channel}.xlsx'
                            file_path = os.path.join(base_dir, file_name)
                            # 通道6文件单独管理
                            if file_path not in [path for _, path in channel_wbs.values()]:
                                wb = Workbook()
                                ws = wb.active
                                ws.append(headers)  # 只写传感器头部，不含OCR和日志
                                channel_wbs[file_path] = (wb, file_path)
                            print(f"\n切换到通道6（来自通道{parent_channel}），数据将保存至 {file_path}")
                            
                        elif cmd[0] == 'log':
                            self.last_log = cmd[1]
                    else:
                        self.ser_com3.write(str(cmd).encode())
                
                # 数据采集逻辑保持不变
                com3_data = self.read_com3_data()
                com6_data = self.read_com6_data()
                com7_data = self.read_com7_data()
                com15_data = self.read_com15_data()
                with self.data_lock:
                    if com3_data and len(com3_data) == 2:
                        self.latest_com3_data = com3_data
                    if com6_data:
                        self.latest_com6_data = com6_data
                    self.latest_com7_data = com7_data
                    self.latest_com15_data = com15_data
                
                # 构建传感器数据行
                merged_data = [
                    *self.latest_com3_data,
                    self.latest_com6_data[0],
                    self.latest_com6_data[1],
                    self.latest_com7_data[0],
                    self.latest_com7_data[1],
                    self.latest_com15_data[0],
                    self.latest_com15_data[1]
                ]
                
                # 应用卡尔曼滤波
                filtered_data = self.apply_kalman_filter(merged_data)
                
                # 添加到缓冲区
                buffer_full = self.add_to_buffer(filtered_data)
                
                # 如果缓冲区已满，处理数据
                if buffer_full:
                    # 处理缓冲区数据（去极值平均）
                    processed_data = self.process_buffer()
                    
                    # 执行OCR识别
                    ocr_results = self.perform_ocr_and_get_data()
                    ocr_values = []
                    if self.ocr_processor.ocr_regions:
                        for region in self.ocr_processor.ocr_regions:
                            region_name = region.get("name", "")
                            if region_name:
                                ocr_values.append(ocr_results.get(region_name, ""))
                    
                    
                    # 新增：读取起始时间文件
                    time_start_file = os.path.join('time_start.txt')
                    if os.path.exists(time_start_file):
                        with open(time_start_file, 'r') as f:
                            start_time_iso = f.read().strip()
                            start_time = datetime.fromisoformat(start_time_iso)
                            print(f"\nstart_time {start_time}")
                    else:
                        start_time = datetime.now()
                        with open(time_start_file, 'w') as f:
                            f.write(start_time.isoformat(timespec='milliseconds'))
                    
                    # 新增：初始化最后一个时间戳
                    
    
                    
                    current_last_timestamp = self.get_last_timestamp_value(main_file_path)
            
                    # 如果文件中有值，使用文件中的值；否则使用内存中的值
                    if current_last_timestamp is not None:
                        last_timestamp_value = current_last_timestamp
            
                    # 计算当前时间与开始时间的差值（秒）
                    current_time = datetime.now()
                    time_diff = (current_time - start_time).total_seconds()
            
                    # 计算需要等待的时间
                    required_diff = last_timestamp_value + 20.0
                    wait_time = max(0, required_diff - time_diff)
                    print(f"\nwait_time {wait_time}")
                    if wait_time > 20:
                        # 等待至满足20秒间隔
                        end_wait = time.time() + 20
                        while time.time() < end_wait and not stop_event.is_set():
                            time.sleep(0.1)
                        if stop_event.is_set():
                            break
                    if 0 < wait_time <= 20:
                        # 等待至满足20秒间隔
                        end_wait = time.time() + wait_time
                        while time.time() < end_wait and not stop_event.is_set():
                            time.sleep(0.1)
                        if stop_event.is_set():
                            break
                            
                        # 更新当前时间
                        current_time = datetime.now()
                        time_diff = (current_time - start_time).total_seconds()
                    
                    # 设置时间戳为上一个值+20秒
                    timestamp_value = last_timestamp_value + 20.0
                    last_timestamp_value = timestamp_value
                    print(f"\nlast_timestamp_value {last_timestamp_value}")
                    
                    
                    # 合并所有数据（时间戳 + 处理后的传感器数据 + OCR数据）
                    #timestamp = datetime.now().isoformat(timespec='milliseconds')
                    all_data = [timestamp_value] + processed_data + ocr_values
                    
                    # 预处理所有数据，将0值替换为0.0001
                    preprocessed_data = self.preprocess_all_data(all_data)
                    #print(f'\r已保存: {preprocessed_data}')
                    # 添加日志信息####################
                    row_merged = preprocessed_data + [self.last_log]
                    #print(f'\r已保存: {row_merged}')
                    if current_channel is not None:
                        # 普通通道数据不含OCR和日志
                        row_channel = preprocessed_data[:len(headers)]
                        buffer.append(row_channel)
                    merged_buffer.append(row_merged)
                    self.last_log = ""
                    
                    try:
                        data_queue.put(row_merged.copy(), block=False, timeout=0.1)
                    except queue.Full:
                        pass
                    n += 1
                    
                    # 写入逻辑优化
                    if len(buffer) >= 10:
                        if current_channel in channel_wbs:
                            wb, _ = channel_wbs[current_channel]
                            ws = wb.active
                            for row in buffer:
                                ws.append(row)
                            buffer.clear()
                    
                    if len(merged_buffer) >= 10:
                        for row in merged_buffer:
                            merged_ws.append(row)
                        merged_buffer.clear()
                        print(f'\r已保存: {row_merged} (总计: {n})', end='', flush=True)
                    #    time.sleep(0.01)
                        
                    # 定期保存工作簿
                    #if time.time() - last_save > 30:  # 每30秒保存一次
                    #    merged_wb.save(main_file_path)
                    #    for key, (wb, path) in channel_wbs.items():
                    #        wb.save(path)
                    #    last_save = time.time()
                    
            pr.disable()
            s = io.StringIO()
            stats = pstats.Stats(pr, stream=s).sort_stats('cumulative')
            stats.print_stats(20)
            print("\n性能分析结果:\n", s.getvalue())
            
        finally:
            # 最终保存所有工作簿
            merged_wb.save(main_file_path)
            for key, (wb, path) in channel_wbs.items():
                wb.save(path)
            
            # 关闭串口连接
            self.ser_com3.close()
            self.instrument_com6.serial.close()
            self.ser_com7.close()
            if self.ser_com15:
                self.ser_com15.close()

                
    def get_last_timestamp_value(self, file_path):
        """从data.xlsx中获取最后一个时间戳值"""
        if not os.path.exists(file_path):
            return 0.0
        
        try:
            # 使用openpyxl读取Excel文件
            wb = load_workbook(file_path)
            ws = wb.active
            
            # 找到最后一行
            last_row = ws.max_row
            
            # 从最后一行开始向上查找最后一个非空的时间戳值
            for row in range(last_row, 1, -1):  # 从最后一行向上搜索
                timestamp_cell = ws.cell(row=row, column=1).value
                if timestamp_cell is not None:
                    return float(timestamp_cell)
            
            # 如果没有找到有效值，返回None
            return 0.0
            
        except Exception as e:
            print(f"读取时间戳值失败: {e}")
            return 0.0

# ------------------ 绘图模块 -------------------
def plot_data():
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(6, 4))
    lines = [
        ax1.plot([], [], 'g-', label='OD')[0],
        ax2.plot([], [], 'b-', label='Fluorescence')[0],
        ax2.plot([], [], 'r-', label='Humidity')[0]
    ]
    timestamps, od_values, fluo_values, hum_values = [], [], [], []
    
    def init():
        for ax in [ax1, ax2]:
            ax.set_xlim(datetime.now(), datetime.now() + timedelta(minutes=10))
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))
        ax1.set_ylabel('OD Value')
        ax2.set_ylabel('Fluorescence/Humidity')
        ax2.legend(loc='upper right')
        return lines
    
    def update(frame):
        nonlocal od_values, fluo_values, hum_values
        if stop_event.is_set():
            plt.close('all')
            return
        while not data_queue.empty():
            data = data_queue.get()
            try:
                ts = datetime.fromisoformat(data[0])
                timestamps.append(ts)
                od_values.append(data[2])
                fluo_values.append(data[3])
                hum_values.append(data[5])
            except Exception as e:
                print(f"绘图数据错误: {e}")
        
        lines[0].set_data(timestamps, od_values)
        lines[1].set_data(timestamps, fluo_values)
        lines[2].set_data(timestamps, hum_values)
        
        for line in lines:
            line.axes.relim()
            line.axes.autoscale_view()
        
        return lines
    
    ani = FuncAnimation(fig, update, init_func=init, blit=True, interval=500)
    plt.tight_layout()
    plt.show()

# ------------------ 主程序 -------------------
if __name__ == "__main__":
    # 打印可用串口列表
    print("可用串口列表:")
    ports = serial.tools.list_ports.comports()
    for port in ports:
        print(f" - {port.device}")
    
    def main():
        controller = MainController()
        collector = DataCollector()
        
        controller_thread = threading.Thread(target=controller.main_loop)
        data_thread = threading.Thread(target=collector.serial_to_excel)
        
        try:
            controller_thread.start()
            data_thread.start()
            
            plot_data()
            
            while not stop_event.is_set():
                if keyboard.is_pressed('esc'):
                    print("\n检测到ESC按键，正在停止...")
                    stop_event.set()
                    
                time.sleep(0.01)
                
        except KeyboardInterrupt:
            print("\n检测到Ctrl+C，正在停止...")
            stop_event.set()
        finally:
            controller_thread.join(timeout=2)
            data_thread.join(timeout=2)
            NewPumpController.pump1.safe_shutdown()
            NewPumpController.pump2.safe_shutdown()
            print("\n程序已安全退出")

    cProfile.runctx('main()', globals(), locals(), filename='performance.prof')