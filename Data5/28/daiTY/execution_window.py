"""
执行窗口模块
显示执行日志和控制执行过程
"""
import time
import threading
import pyautogui
import pygetwindow as gw
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
                             QPushButton, QTextEdit, QGroupBox, QLineEdit, 
                             QCheckBox, QProgressBar)
from PyQt5.QtGui import QIcon, QFont
from PyQt5.QtCore import pyqtSignal, Qt, QObject
import cv2
import numpy as np
from PIL import ImageGrab

import os
import sqlite3
import csv
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from config_manager import set_last_operation, load_config

class ExecutionWindow(QWidget):
    log_signal = pyqtSignal(str, str)
    progress_signal = pyqtSignal(int)
    enable_btn_signal = pyqtSignal(bool, bool, bool)
    
    def __init__(self, config_name):
        super().__init__()
        self.config_name = config_name
        self.running = False
        self.stop_flag = False
        self.setup_ui()
        self.load_config()
        set_last_operation(self.config_name)
        
        self.log_signal.connect(self.log)
        self.progress_signal.connect(self.progress_bar.setValue)
        self.enable_btn_signal.connect(self.set_btn_state)
    
    def setup_ui(self):
        """初始化UI界面"""
        self.setWindowTitle(f"执行配置 - {self.config_name}")
        self.setGeometry(100, 100, 800, 600)
        self.setMinimumSize(700, 500)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # 标题
        title_label = QLabel(f"执行配置: {self.config_name}")
        title_label.setFont(QFont("微软雅黑", 16, QFont.Bold))
        layout.addWidget(title_label)
        
        # 配置信息组
        config_group = QGroupBox("配置信息")
        config_group.setFont(QFont("微软雅黑", 13, QFont.Bold))
        config_layout = QVBoxLayout(config_group)
        self.config_info = QLabel()
        self.config_info.setFont(QFont("微软雅黑", 12))
        config_layout.addWidget(self.config_info)
        layout.addWidget(config_group)
        
        # 执行控制组
        control_group = QGroupBox("执行控制")
        control_group.setFont(QFont("微软雅黑", 13, QFont.Bold))
        control_layout = QHBoxLayout(control_group)
        
        self.start_btn = QPushButton("开始执行")
        self.start_btn.setFont(QFont("微软雅黑", 13))
        self.start_btn.setIcon(QIcon("icons/start.png"))
        self.start_btn.clicked.connect(self.start_operation)
        control_layout.addWidget(self.start_btn)
        
        self.pause_btn = QPushButton("暂停")
        self.pause_btn.setFont(QFont("微软雅黑", 13))
        self.pause_btn.setIcon(QIcon("icons/pause.png"))
        self.pause_btn.setEnabled(False)
        self.pause_btn.clicked.connect(self.pause_operation)
        control_layout.addWidget(self.pause_btn)
        
        self.stop_btn = QPushButton("停止")
        self.stop_btn.setFont(QFont("微软雅黑", 13))
        self.stop_btn.setIcon(QIcon("icons/stop.png"))
        self.stop_btn.setEnabled(False)
        self.stop_btn.clicked.connect(self.stop_operation)
        control_layout.addWidget(self.stop_btn)
        
        layout.addWidget(control_group)
        
        # 循环控制
        loop_layout = QHBoxLayout()
        self.loop_check = QCheckBox("循环执行")
        self.loop_check.setFont(QFont("微软雅黑", 12))
        loop_layout.addWidget(self.loop_check)
        
        loop_layout.addWidget(QLabel("循环次数:"))
        self.loop_count = QLineEdit("1")
        self.loop_count.setFont(QFont("微软雅黑", 12))
        self.loop_count.setFixedWidth(50)
        loop_layout.addWidget(self.loop_count)
        
        loop_layout.addWidget(QLabel("间隔(秒):"))
        self.loop_interval = QLineEdit("1.0")
        self.loop_interval.setFont(QFont("微软雅黑", 12))
        self.loop_interval.setFixedWidth(50)
        loop_layout.addWidget(self.loop_interval)
        
        loop_layout.addStretch()
        layout.addLayout(loop_layout)
        
        # 进度和日志组
        progress_group = QGroupBox("执行进度和日志")
        progress_group.setFont(QFont("微软雅黑", 13, QFont.Bold))
        progress_layout = QVBoxLayout(progress_group)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setFont(QFont("微软雅黑", 11))
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        progress_layout.addWidget(self.progress_bar)
        
        self.log_output = QTextEdit()
        self.log_output.setReadOnly(True)
        self.log_output.setFont(QFont("Consolas", 10))
        progress_layout.addWidget(self.log_output)
        
        layout.addWidget(progress_group)
        
        # 底部按钮
        bottom_layout = QHBoxLayout()
        bottom_layout.addStretch()
        
        self.close_btn = QPushButton("关闭")
        self.close_btn.setFont(QFont("微软雅黑", 13))
        self.close_btn.setIcon(QIcon("icons/close.png"))
        self.close_btn.clicked.connect(self.close)
        bottom_layout.addWidget(self.close_btn)
        
        layout.addLayout(bottom_layout)
    
    def load_config(self):
        """加载配置详情"""
        # 如果外部已传入config（如测试模式），则直接用
        if hasattr(self, "config") and self.config:
            config = self.config
        else:
            config = load_config(self.config_name)
            if not config:
                self.config = {
                    "window_title": "",
                    "operations": [],
                    "loop_times": 1,
                    "loop_interval": 1.0
                }
                self.config_info.setText("未找到配置文件")
                return
        self.config = config

        # 修正：确保 operations 为列表
        operations = self.config.get('operations')
        if operations is None:
            operations = []
            self.config['operations'] = operations

        info_text = f"窗口标题: {self.config.get('window_title', '无')}\n"
        info_text += f"操作步骤数量: {len(operations)}\n"
        info_text += f"循环次数: {self.config.get('loop_times', 1)}\n"
        info_text += f"循环间隔: {self.config.get('loop_interval', 0)}秒"
        self.config_info.setText(info_text)
        self.progress_bar.setMaximum(len(operations))
    
    def set_btn_state(self, start, pause, stop):
        """设置按钮状态"""
        self.start_btn.setEnabled(start)
        self.pause_btn.setEnabled(pause)
        self.stop_btn.setEnabled(stop)

    def log(self, message, level="INFO"):
        """记录日志（主线程安全）"""
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
        log_entry = f"[{timestamp}] [{level}] {message}"
        self.log_output.append(log_entry)
        self.log_output.verticalScrollBar().setValue(
            self.log_output.verticalScrollBar().maximum()
        )
    
    def start_operation(self):
        """开始执行操作"""
        if self.running:
            return
            
        self.log("开始执行操作...")
        self.running = True
        self.stop_flag = False
        
        # 更新按钮状态
        self.start_btn.setEnabled(False)
        self.pause_btn.setEnabled(True)
        self.stop_btn.setEnabled(True)
        
        # 启动操作线程
        self.operation_thread = threading.Thread(target=self.run_operations)
        self.operation_thread.daemon = True
        self.operation_thread.start()
    
    def pause_operation(self):
        """暂停操作"""
        self.log("暂停功能尚未实现", "WARNING")
        self.pause_btn.setEnabled(False)
    
    def stop_operation(self):
        """停止操作"""
        if not self.running:
            return
            
        self.log("正在停止操作...")
        self.stop_flag = True
        self.running = False
        self.stop_btn.setEnabled(False)
        self.pause_btn.setEnabled(False)
        self.start_btn.setEnabled(True)
    
    def run_operations(self):
        """执行操作步骤"""
        try:
            window_title = self.config.get("window_title", "")
            if not window_title:
                self.log_signal.emit("未指定目标窗口标题", "ERROR")
                self.stop_operation()
                return

            # 获取操作步骤
            operations = self.config.get("operations", [])
            if not operations or not isinstance(operations, list):
                self.log_signal.emit("未配置任何操作步骤", "ERROR")
                self.stop_operation()
                return
            total_steps = len(operations)

            # 查找窗口
            windows = gw.getWindowsWithTitle(window_title)
            if not windows:
                self.log_signal.emit(f"未找到窗口: {window_title}", "ERROR")
                self.stop_operation()
                return
            target_window = windows[0]
            wx, wy, wwidth, wheight = target_window.left, target_window.top, target_window.width, target_window.height

            # 数据源处理
            data_source = self.config.get("data_source")
            excel_path = None
            if data_source and data_source.get("type") == "excel":
                excel_path = data_source["path"]
            
            # 如果没有数据源，使用普通执行模式
            if not excel_path:
                self.log_signal.emit("没有配置Excel数据源，使用普通执行模式", "WARNING")
                return self.run_normal_operations(target_window, operations, wx, wy)
            
            # Excel数据驱动执行
            return self.run_excel_driven_operations(target_window, operations, excel_path, wx, wy)
            
        except Exception as e:
            self.log_signal.emit(f"执行过程中发生错误: {str(e)}", "ERROR")
        finally:
            self.running = False
            self.stop_flag = False
            self.enable_btn_signal.emit(True, False, False)

    def run_normal_operations(self, window, operations, wx, wy):
        """普通执行模式（无数据驱动）"""
        try:
            # 每次循环都确保窗口在最前并激活
            try:
                window.activate()
                window.restore()
                window.bringToFront()
            except Exception:
                pass
            time.sleep(0.2)

            # 执行每个步骤
            for index, operation in enumerate(operations):
                if self.stop_flag:
                    break
                self.progress_signal.emit(index + 1)
                self.execute_operation(window, operation, index, wx, wy)

            if not self.stop_flag:
                self.log_signal.emit("所有操作执行完成！", "INFO")
                self.progress_signal.emit(len(operations))
        except Exception as e:
            self.log_signal.emit(f"普通执行模式出错: {str(e)}", "ERROR")

    def run_excel_driven_operations(self, window, operations, excel_path, wx, wy):
        """Excel数据驱动执行模式（基于时间戳判断更新）"""
        try:
            # 获取标签与操作的映射关系
            tag_to_operation = {}
            for op in operations:
                region = op.get("region", {})
                tag = region.get("name", "").replace("Excel标签: ", "")
                if tag:
                    tag_to_operation[tag] = op
            
            # 检查是否有有效的映射关系
            if not tag_to_operation:
                self.log_signal.emit("未找到有效的标签映射关系", "ERROR")
                return
            
            self.log_signal.emit(f"已建立标签映射关系: {', '.join(tag_to_operation.keys())}", "INFO")
            
            # 初始化文件监控状态
            last_modified_time = 0  # 最后修改时间戳
            last_row_count = 0      # 最后行数
            timeout = 20 * 60       # 20分钟超时
            last_update_time = time.time()  # 最后更新时间
            
            while not self.stop_flag:
                # 检查文件是否存在
                if not os.path.exists(excel_path):
                    self.log_signal.emit(f"Excel文件不存在: {excel_path}", "ERROR")
                    time.sleep(10)
                    continue
                
                try:
                    # 获取当前文件修改时间
                    current_modified_time = os.path.getmtime(excel_path)
                    
                    # 检查文件是否有更新
                    if current_modified_time > last_modified_time:
                        self.log_signal.emit(f"检测到文件更新 (时间戳: {current_modified_time})", "INFO")
                        
                        # 读取Excel文件
                        wb = load_workbook(excel_path, read_only=True)
                        ws = wb.active
                        
                        # 获取标题行（第一行）
                        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                        
                        # 检查是否有数据行
                        row_count = ws.max_row
                        if row_count <= 1:
                            self.log_signal.emit("Excel文件中无数据行", "WARNING")
                            wb.close()
                            # 更新最后修改时间（避免重复提示）
                            last_modified_time = current_modified_time
                            time.sleep(10)
                            continue
                        
                        # 获取最后一行数据
                        last_data_row = next(ws.iter_rows(min_row=row_count, max_row=row_count, values_only=True))
                        
                        # 更新状态
                        last_modified_time = current_modified_time
                        last_row_count = row_count
                        last_update_time = time.time()
                        
                        # 确保窗口激活
                        try:
                            window.activate()
                            window.restore()
                        except Exception:
                            pass
                        time.sleep(0.5)
                        
                        # 执行操作
                        self.log_signal.emit("开始执行操作...", "INFO")
                        for col_idx, value in enumerate(last_data_row):
                            if col_idx == 0:  # 跳过第一列（时间戳）
                                continue
                            
                            if col_idx - 1 >= len(header_row):
                                break
                            
                            tag = header_row[col_idx]  # 获取标签
                            op = tag_to_operation.get(tag)
                            
                            if not op:
                                self.log_signal.emit(f"未找到标签'{tag}'对应的操作", "WARNING")
                                continue
                            
                            # 更新操作的文本值
                            if "config" not in op:
                                op["config"] = {}
                            op["config"]["text"] = str(value)
                            
                            # 执行操作
                            self.execute_operation(window, op, col_idx - 1, wx, wy)
                        
                        self.log_signal.emit("操作执行完成！", "INFO")
                        self.progress_signal.emit(len(tag_to_operation))
                    else:
                        # 检查超时
                        if time.time() - last_update_time > timeout:
                            self.log_signal.emit("20分钟内无新数据，退出执行", "INFO")
                            break
                        
                        # 计算等待时间
                        wait_seconds = int(time.time() - last_update_time)
                        self.log_signal.emit(f"等待新数据... (已等待 {wait_seconds}秒)", "INFO")
                    
                    wb.close()
                except Exception as e:
                    self.log_signal.emit(f"处理Excel文件出错: {str(e)}", "ERROR")
                
                # 等待10秒后再次检查
                for i in range(10):
                    if self.stop_flag:
                        break
                    time.sleep(1)
            
            self.log_signal.emit("Excel数据驱动执行结束", "INFO")
        except Exception as e:
            self.log_signal.emit(f"Excel数据驱动执行出错: {str(e)}", "ERROR")

    def execute_operation(self, window, operation, step_index, wx=None, wy=None, wwidth=None, wheight=None):
        """执行单个操作，支持OpenCV模板匹配，并自适应多屏坐标"""
        op_type = operation.get("action") or operation.get("type")
        region = operation.get("region")
        text = operation.get("text", "") or operation.get("config", {}).get("text", "")
        wait_time = operation.get("wait_time", 0.5)
        times = operation.get("config", {}).get("times", 1)

        if not region or not isinstance(region, dict):
            self.log_signal.emit(f"步骤 {step_index+1}: 区域信息无效，跳过", "WARNING")
            return

        template_path = region.get("template_path")
        found = False

        # 优先模板匹配
        if template_path and os.path.exists(template_path):
            try:
                bbox = (window.left, window.top, window.left + window.width, window.top + window.height)
                screenshot = ImageGrab.grab(bbox)
                screenshot_np = np.array(screenshot)
                screenshot_gray = cv2.cvtColor(screenshot_np, cv2.COLOR_BGR2GRAY)
                template = cv2.imread(template_path, 0)
                res = cv2.matchTemplate(screenshot_gray, template, cv2.TM_CCOEFF_NORMED)
                min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(res)
                threshold = 0.85
                if max_val >= threshold:
                    t_height, t_width = template.shape
                    center_x = window.left + max_loc[0] + t_width // 2
                    center_y = window.top + max_loc[1] + t_height // 2
                    found = True
                    self.log_signal.emit(f"步骤 {step_index+1}: 模板匹配成功，点击({center_x},{center_y})", "INFO")
                    if op_type == "click":
                        for t in range(times):
                            pyautogui.click(x=center_x, y=center_y)
                            time.sleep(wait_time)
                    elif op_type == "input":
                        pyautogui.click(x=center_x, y=center_y)
                        pyautogui.write(text)
                        time.sleep(wait_time)
                    elif op_type == "wait":
                        time.sleep(wait_time)
                    return
            except Exception as e:
                self.log_signal.emit(f"OpenCV模板匹配失败: {e}", "WARNING")

        if not found:
            # 回退到原有坐标点击（多屏自适应）
            if wx is None or wy is None:
                wx, wy = window.left, window.top
            rect = region.get("rect", region)
            # 修正：x/y 应该是窗口内的相对坐标，不能直接用全局坐标
            x = rect.get("x", 0)
            y = rect.get("y", 0)
            width = rect.get("width", 0)
            height = rect.get("height", 0)
            center_x = wx + x + width // 2
            center_y = wy + y + height // 2

            self.log_signal.emit(
                f"DEBUG: wx={wx}, wy={wy}, x={x}, y={y}, width={width}, height={height}, center=({center_x},{center_y})", "DEBUG"
            )

            try:
                window.activate()
                window.restore()
            except Exception:
                pass
            time.sleep(0.1)

            if op_type == "click":
                for t in range(times):
                    self.log_signal.emit(f"步骤 {step_index+1}: 点击位置 ({center_x}, {center_y}) 第{t+1}次", "INFO")
                    pyautogui.click(x=center_x, y=center_y)
                    time.sleep(wait_time)
            elif op_type == "input":
                self.log_signal.emit(f"步骤 {step_index+1}: 在位置 ({center_x}, {center_y}) 输入文本: {text}", "INFO")
                pyautogui.click(x=center_x, y=center_y)
                pyautogui.write(text)
                time.sleep(wait_time)
            elif op_type == "wait":
                self.log_signal.emit(f"步骤 {step_index+1}: 等待 {wait_time} 秒", "INFO")
                time.sleep(wait_time)
                return
            else:
                self.log_signal.emit(f"步骤 {step_index+1}: 未知操作类型 {op_type}，跳过", "WARNING")
    
    def closeEvent(self, event):
        """处理窗口关闭事件"""
        self.stop_operation()
        event.accept()